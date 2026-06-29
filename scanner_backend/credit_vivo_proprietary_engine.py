from __future__ import annotations

"""
Credit Vivo Proprietary Parser Engine v16

Purpose:
- Build a stronger original Credit Vivo parser/scanner engine.
- No Claude / Anthropic / paid AI API.
- No competitor code.
- No automatic dispute sending.
- Every output includes source evidence and confidence.

Core ideas:
1. Extract all raw report text first.
2. Detect bureau and report layout.
3. Segment report into account/profile blocks.
4. Extract fields into Credit Vivo's normalized tradeline schema.
5. Score confidence.
6. Match same/similar accounts across bureaus.
7. Detect review issues.
8. Return customer-friendly and admin-ready output.

This is draft review software. It does not provide legal advice.
"""

from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Tuple, Iterable
import hashlib
import re
import json
import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


# -----------------------------
# Utility
# -----------------------------

def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def stable_id(*parts: str) -> str:
    raw = "|".join(parts).encode("utf-8", errors="ignore")
    return "cvp_" + hashlib.sha256(raw).hexdigest()[:16]


def money_to_number(value: str) -> Optional[float]:
    if not value:
        return None
    value = value.replace("$", "").replace(",", "").strip()
    try:
        return float(value)
    except Exception:
        return None


def normalize_money(value: str) -> str:
    n = money_to_number(value)
    if n is None:
        return value.strip()
    if n.is_integer():
        return f"${int(n):,}"
    return f"${n:,.2f}"


def normalize_date(value: str) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value.strip())


def mask_account_number(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    digits = re.sub(r"\D", "", value)
    if len(digits) >= 4:
        return "*" + digits[-4:]
    if len(value) >= 6:
        return value[:2] + "..." + value[-2:]
    return value


def simple_similarity(a: str, b: str) -> float:
    """
    Custom no-dependency similarity for proprietary matching.
    Not perfect, but good enough for candidate grouping without RapidFuzz.
    """
    ak = compact_key(a)
    bk = compact_key(b)
    if not ak or not bk:
        return 0.0
    if ak == bk:
        return 1.0
    if ak in bk or bk in ak:
        return 0.86

    aset = set(re.findall(r"[a-z0-9]{2,}", a.lower()))
    bset = set(re.findall(r"[a-z0-9]{2,}", b.lower()))
    if not aset or not bset:
        return 0.0
    return len(aset & bset) / max(1, len(aset | bset))


# -----------------------------
# Schema
# -----------------------------

@dataclass
class Evidence:
    bureau: str
    page: Optional[int]
    snippet: str
    extraction_method: str = "native_rule_engine"


@dataclass
class NormalizedTradeline:
    id: str
    bureau: str
    source_filename: str
    account_name: str = ""
    account_number_masked: str = ""
    account_type: str = ""
    portfolio_type: str = ""
    responsibility: str = ""
    creditor_classification: str = ""
    original_creditor: str = ""
    collector_or_debt_buyer: str = ""
    status: str = ""
    pay_status: str = ""
    balance: str = ""
    past_due: str = ""
    high_credit_or_original_amount: str = ""
    credit_limit: str = ""
    date_opened: str = ""
    date_closed: str = ""
    date_reported: str = ""
    date_last_activity: str = ""
    date_last_payment: str = ""
    date_of_first_delinquency: str = ""
    estimated_removal_date: str = ""
    remarks: str = ""
    payment_history_summary: str = ""
    raw_block: str = ""
    page_start: Optional[int] = None
    confidence: str = "medium"
    confidence_score: float = 0.0
    needs_admin_review: bool = True


@dataclass
class ReviewIssue:
    id: str
    issue_type: str
    severity: str
    customer_label: str
    customer_explanation: str
    admin_explanation: str
    suggested_round: str
    related_tradeline_ids: List[str] = field(default_factory=list)
    evidence: List[Evidence] = field(default_factory=list)
    confidence: str = "medium"


@dataclass
class ParseResult:
    engine: str
    version: str
    paid_ai_used: bool
    files: List[dict]
    tradelines: List[NormalizedTradeline]
    issues: List[ReviewIssue]
    cross_bureau_groups: List[dict]
    customer_summary: dict
    admin_summary: dict


# -----------------------------
# Bureau profiles
# -----------------------------

DATE = r"(?:\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{4})"
MONEY = r"\$?\s?[0-9]{1,3}(?:,[0-9]{3})*(?:\.\d{2})?|\$?\s?[0-9]+(?:\.\d{2})?"

COMMON_FIELD_PATTERNS: Dict[str, List[str]] = {
    "account_number_masked": [
        r"(?:account\s*(?:#|number|no\.?)|acct\s*(?:#|number|no\.?))\s*[:\-]?\s*([A-Za-z0-9\*\-xX]{3,32})",
    ],
    "account_type": [
        r"(?:account type|type of account|loan type|type)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,60})",
    ],
    "portfolio_type": [
        r"(?:portfolio type|portfolio)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,60})",
    ],
    "responsibility": [
        r"(?:responsibility|account holder|owner)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,60})",
    ],
    "status": [
        r"(?:account status|status)\s*[:\-]?\s*([A-Za-z0-9 /&.,'-]{3,100})",
    ],
    "pay_status": [
        r"(?:payment status|pay status|payment\s*condition)\s*[:\-]?\s*([A-Za-z0-9 /&.,'-]{3,120})",
    ],
    "balance": [
        r"(?:current balance|balance)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "past_due": [
        r"(?:past due|amount past due)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "high_credit_or_original_amount": [
        r"(?:high credit|original amount|original balance|loan amount)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "credit_limit": [
        r"(?:credit limit|limit)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "date_opened": [
        r"(?:date opened|opened)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_closed": [
        r"(?:date closed|closed)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_reported": [
        r"(?:date reported|last reported|reported|updated|date updated)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_last_activity": [
        r"(?:date of last activity|last activity|date last active)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_last_payment": [
        r"(?:date of last payment|last payment|date last payment)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_of_first_delinquency": [
        r"(?:date of first delinquency|first delinquency|dofd|first reported delinquency)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "estimated_removal_date": [
        r"(?:on record until|estimated month and year this item will be removed|estimated removal|scheduled to continue on record until)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "original_creditor": [
        r"(?:original creditor|original lender|original account)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})",
    ],
    "collector_or_debt_buyer": [
        r"(?:collection agency|collector|debt buyer|assigned to)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})",
    ],
    "creditor_classification": [
        r"(?:creditor classification|classification|industry type)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,80})",
    ],
    "remarks": [
        r"(?:remarks|comments|comment|account information)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{5,220})",
    ],
}

BUREAU_SIGNATURES = {
    "Experian": ["experian", "experian credit report", "experian information solutions"],
    "Equifax": ["equifax", "equifax information services", "econsumer"],
    "TransUnion": ["transunion", "trans union", "transunion llc"],
}

NEGATIVE_TERMS = [
    "collection", "charge off", "charge-off", "charged off", "past due", "late",
    "delinquent", "derogatory", "settled", "repossession", "foreclosure",
    "120 days", "90 days", "60 days", "30 days", "placed for collection",
    "transferred", "sold", "profit and loss", "written off", "bad debt",
    "unpaid", "seriously past due"
]

COLLECTION_TERMS = [
    "collection", "collection agency", "collector", "debt buyer",
    "original creditor", "placed for collection", "assigned"
]

BOILERPLATE_TERMS = [
    "consumerfinance.gov",
    "violates the fcra",
    "additional information",
    "the last reported status of the account",
    "name, address, and phone",
    "insurer, employer, landlord",
    "upgrades and enhancements",
    "account types is good for your credit",
    "amount of the debt",
    "public records and residential information",
    "supplemental public records",
]

BAD_ACCOUNT_NAME_FRAGMENTS = [
    "consumerfinance.gov",
    "learnmore",
    "violates the fcra",
    "additional information",
    "last reported status",
    "account types is good",
    "amount of the debt",
    "name, address",
    "insurer, employer",
    "landlord",
    "enhancements",
    "responsibility individual",
    "public records and residential",
    "supplemental public records",
]


# -----------------------------
# Text + page utilities
# -----------------------------

def detect_bureau(filename: str, text: str) -> str:
    sample = (filename + "\n" + text[:5000]).lower()
    scores = {}
    for bureau, terms in BUREAU_SIGNATURES.items():
        scores[bureau] = sum(1 for t in terms if t in sample)
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "Unknown Bureau"


def page_split(text: str) -> List[Tuple[Optional[int], str]]:
    """
    Supports extraction format with --- PAGE 1 --- markers.
    """
    parts = re.split(r"\n?\s*---\s*PAGE\s+(\d+)\s*---\s*\n?", text, flags=re.I)
    if len(parts) <= 1:
        return [(None, text)]
    pages = []
    preface = parts[0]
    for i in range(1, len(parts), 2):
        try:
            page_num = int(parts[i])
        except Exception:
            page_num = None
        page_text = parts[i + 1] if i + 1 < len(parts) else ""
        pages.append((page_num, page_text))
    return pages


def candidate_blocks(text: str) -> List[Tuple[Optional[int], str]]:
    blocks: List[Tuple[Optional[int], str]] = []

    for page_num, page_text in page_split(text):
        page_text = clean_text(page_text)
        if not page_text:
            continue

        # Main split using paragraph gaps
        chunks = re.split(r"\n\s*\n", page_text)
        buffer: List[str] = []

        for chunk in chunks:
            c = clean_text(chunk)
            if len(c) < 20:
                continue
            lower = c.lower()
            starts = (
                any(label in lower for label in ["account number", "account #", "date opened", "payment status", "account status"])
                or any(term in lower for term in NEGATIVE_TERMS)
            )
            if starts and buffer:
                block = "\n".join(buffer)
                if len(block) > 80:
                    blocks.append((page_num, block))
                buffer = [c]
            else:
                buffer.append(c)

            if len("\n".join(buffer)) > 3500:
                blocks.append((page_num, "\n".join(buffer)))
                buffer = []

        if buffer:
            block = "\n".join(buffer)
            if len(block) > 80:
                blocks.append((page_num, block))

        # Secondary sliding windows for negative terms
        lower_page = page_text.lower()
        for term in NEGATIVE_TERMS:
            for m in re.finditer(re.escape(term), lower_page):
                start = max(0, m.start() - 800)
                end = min(len(page_text), m.end() + 1300)
                blocks.append((page_num, page_text[start:end]))

    return dedupe_blocks(blocks)


def dedupe_blocks(blocks: List[Tuple[Optional[int], str]]) -> List[Tuple[Optional[int], str]]:
    seen = set()
    out = []
    for page_num, block in blocks:
        clean = clean_text(block)
        if len(clean) < 80:
            continue
        key = hashlib.sha1(clean[:700].encode("utf-8", errors="ignore")).hexdigest()
        if key not in seen:
            seen.add(key)
            out.append((page_num, clean[:4500]))
    return out


# -----------------------------
# Field extraction
# -----------------------------

def first_match(patterns: List[str], text: str) -> str:
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I | re.M)
        if m:
            val = clean_text(m.group(1))
            val = re.split(r"\n| {3,}", val)[0].strip(" :-")
            return val[:240]
    return ""


def extract_field(field_name: str, block: str) -> str:
    value = first_match(COMMON_FIELD_PATTERNS.get(field_name, []), block)
    if field_name in {"balance", "past_due", "high_credit_or_original_amount", "credit_limit"}:
        return normalize_money(value)
    if "date" in field_name:
        return normalize_date(value)
    if field_name == "account_number_masked":
        return mask_account_number(value)
    return value


def is_bad_account_name(name: str) -> bool:
    cleaned = clean_text(name).strip(" .;:,|-")
    lower = cleaned.lower()
    if not cleaned or cleaned == "Review Item":
        return True
    if len(cleaned) < 3 or len(cleaned) > 90:
        return True
    if any(fragment in lower for fragment in BAD_ACCOUNT_NAME_FRAGMENTS):
        return True
    if lower.startswith(("of ", "and ", "the ", "this ", "that ", "www.")):
        return True
    if cleaned.endswith(".") and len(cleaned.split()) > 3:
        return True
    if len(cleaned.split()) > 8:
        return True
    if not re.search(r"[A-Za-z]", cleaned):
        return True
    return False


def guess_account_name(block: str) -> str:
    lines = [clean_text(x).strip(" :-") for x in block.splitlines() if clean_text(x)]
    bad = (
        "account number", "account #", "balance", "date ", "status", "payment",
        "remarks", "comments", "address", "phone", "page ", "experian", "equifax",
        "transunion", "trans union", "credit report", "report number", "summary"
    )

    labeled = first_match([
        r"(?:account name|creditor|furnisher|subscriber)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})"
    ], block)
    if labeled:
        return labeled if not is_bad_account_name(labeled) else "Review Item"

    for line in lines[:14]:
        lower = line.lower()
        if any(lower.startswith(x) for x in bad):
            continue
        if len(line) < 3 or len(line) > 85:
            continue
        if not re.search(r"[A-Za-z]", line):
            continue
        # Avoid long report sentences
        if len(line.split()) <= 9 and not is_bad_account_name(line):
            return line

    return "Review Item"


def is_boilerplate_block(block: str) -> bool:
    lower = block.lower()
    boilerplate_hits = sum(1 for term in BOILERPLATE_TERMS if term in lower)
    core_field_hits = sum(
        1
        for term in [
            "account number",
            "account #",
            "date opened",
            "date reported",
            "original creditor",
            "account type",
            "payment status",
            "account status",
            "balance",
        ]
        if term in lower
    )
    return boilerplate_hits > 0 and core_field_hits < 2


def is_probable_tradeline(t: NormalizedTradeline) -> bool:
    if is_bad_account_name(t.account_name):
        return False

    core_fields = [
        t.account_number_masked,
        t.account_type,
        t.status or t.pay_status,
        t.balance,
        t.date_opened,
        t.date_reported,
        t.date_of_first_delinquency,
        t.original_creditor,
        t.remarks,
    ]
    core_count = sum(1 for field in core_fields if field)
    lower = t.raw_block.lower()
    has_money_or_account = bool(t.balance or t.account_number_masked)
    has_credit_signal = any(
        term in lower
        for term in [
            "account type",
            "original creditor",
            "date opened",
            "date reported",
            "payment status",
            "account status",
            "credit limit",
            "past due",
            "high credit",
        ]
    )
    return core_count >= 2 and (has_money_or_account or has_credit_signal)


def classify_category(t: NormalizedTradeline) -> Tuple[str, str, str, str]:
    blob = " ".join([
        t.account_name, t.account_type, t.status, t.pay_status, t.remarks,
        t.original_creditor, t.collector_or_debt_buyer, t.raw_block
    ]).lower()

    if any(x in blob for x in COLLECTION_TERMS):
        return (
            "Collection Review",
            "medium",
            "Collection or debt-buyer information should be reviewed.",
            "Round 2 — Collection Review"
        )

    if "charge" in blob and "off" in blob:
        return (
            "Reporting Accuracy Review",
            "medium",
            "Charge-off reporting should be reviewed for balance, status, dates, and ownership.",
            "Round 4 — Reporting Accuracy Review"
        )

    if any(x in blob for x in ["transferred", "sold", "closed"]):
        return (
            "Reporting Accuracy Review",
            "medium",
            "Transferred, sold, or closed reporting should be reviewed for accuracy.",
            "Round 4 — Reporting Accuracy Review"
        )

    if any(x in blob for x in NEGATIVE_TERMS):
        return (
            "Factual Review",
            "medium",
            "This negative item should be reviewed for factual accuracy.",
            "Round 5 — Factual Review"
        )

    return (
        "Bureau Match Review",
        "low",
        "Compare this item across bureaus for consistency.",
        "Round 3 — Bureau Match Review"
    )


def score_confidence(t: NormalizedTradeline) -> float:
    score = 0.0
    if t.account_name and t.account_name != "Review Item":
        score += 0.22
    if t.account_number_masked:
        score += 0.16
    if t.balance:
        score += 0.10
    if t.status or t.pay_status:
        score += 0.13
    if t.date_opened:
        score += 0.08
    if t.date_reported:
        score += 0.08
    if t.date_of_first_delinquency or t.estimated_removal_date:
        score += 0.10
    if t.raw_block and len(t.raw_block) > 250:
        score += 0.08
    if any(x in t.raw_block.lower() for x in NEGATIVE_TERMS):
        score += 0.05
    return min(1.0, score)


def confidence_label(score: float) -> str:
    if score >= 0.72:
        return "high"
    if score >= 0.42:
        return "medium"
    return "low"


def parse_tradelines_for_bureau(bureau: str, filename: str, text: str) -> List[NormalizedTradeline]:
    tradelines = []
    for page_num, block in candidate_blocks(text):
        lower = block.lower()
        if is_boilerplate_block(block):
            continue
        has_signal = (
            any(term in lower for term in NEGATIVE_TERMS)
            or any(label in lower for label in ["account number", "account #", "date opened", "payment status", "account status"])
        )
        if not has_signal:
            continue

        account_name = guess_account_name(block)
        t = NormalizedTradeline(
            id=stable_id(bureau, filename, account_name, block[:160]),
            bureau=bureau,
            source_filename=filename,
            account_name=account_name,
            account_number_masked=extract_field("account_number_masked", block),
            account_type=extract_field("account_type", block),
            portfolio_type=extract_field("portfolio_type", block),
            responsibility=extract_field("responsibility", block),
            creditor_classification=extract_field("creditor_classification", block),
            original_creditor=extract_field("original_creditor", block),
            collector_or_debt_buyer=extract_field("collector_or_debt_buyer", block),
            status=extract_field("status", block),
            pay_status=extract_field("pay_status", block),
            balance=extract_field("balance", block),
            past_due=extract_field("past_due", block),
            high_credit_or_original_amount=extract_field("high_credit_or_original_amount", block),
            credit_limit=extract_field("credit_limit", block),
            date_opened=extract_field("date_opened", block),
            date_closed=extract_field("date_closed", block),
            date_reported=extract_field("date_reported", block),
            date_last_activity=extract_field("date_last_activity", block),
            date_last_payment=extract_field("date_last_payment", block),
            date_of_first_delinquency=extract_field("date_of_first_delinquency", block),
            estimated_removal_date=extract_field("estimated_removal_date", block),
            remarks=extract_field("remarks", block),
            raw_block=block[:2500],
            page_start=page_num,
        )
        t.confidence_score = score_confidence(t)
        t.confidence = confidence_label(t.confidence_score)
        t.needs_admin_review = t.confidence != "high"
        if not is_probable_tradeline(t):
            continue
        tradelines.append(t)

    return dedupe_tradelines(tradelines)


def dedupe_tradelines(items: List[NormalizedTradeline]) -> List[NormalizedTradeline]:
    out = []
    seen = set()
    for t in items:
        key = (
            t.bureau,
            compact_key(t.account_name)[:35],
            t.account_number_masked,
            t.page_start,
            compact_key(t.balance),
            compact_key(t.status or t.pay_status)[:25],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


# -----------------------------
# Cross-bureau matching
# -----------------------------

def group_cross_bureau(tradelines: List[NormalizedTradeline]) -> List[dict]:
    groups: List[dict] = []
    used = set()

    for i, t in enumerate(tradelines):
        if t.id in used:
            continue

        group = [t]
        used.add(t.id)

        for other in tradelines[i+1:]:
            if other.id in used:
                continue
            if t.bureau == other.bureau:
                continue

            name_score = simple_similarity(t.account_name, other.account_name)
            acct_match = bool(t.account_number_masked and t.account_number_masked == other.account_number_masked)
            original_score = simple_similarity(t.original_creditor, other.original_creditor) if t.original_creditor and other.original_creditor else 0

            if acct_match or name_score >= 0.72 or original_score >= 0.72:
                group.append(other)
                used.add(other.id)

        if len(group) >= 2:
            groups.append({
                "group_id": stable_id("group", *[x.id for x in group]),
                "bureaus": sorted({x.bureau for x in group}),
                "account_names": sorted({x.account_name for x in group}),
                "tradeline_ids": [x.id for x in group],
                "review_note": "Same or similar item appears across multiple bureaus. Compare balance, status, dates, original creditor, and remarks.",
                "field_snapshot": [
                    {
                        "bureau": x.bureau,
                        "account_name": x.account_name,
                        "balance": x.balance,
                        "status": x.status,
                        "pay_status": x.pay_status,
                        "date_opened": x.date_opened,
                        "date_reported": x.date_reported,
                        "date_of_first_delinquency": x.date_of_first_delinquency,
                        "estimated_removal_date": x.estimated_removal_date,
                    }
                    for x in group
                ]
            })

    return groups


# -----------------------------
# Issue detection engine
# -----------------------------

def ev(t: NormalizedTradeline) -> Evidence:
    return Evidence(
        bureau=t.bureau,
        page=t.page_start,
        snippet=t.raw_block[:800],
    )


def add_issue(issues: List[ReviewIssue], issue_type: str, severity: str, label: str,
              customer: str, admin: str, round_name: str, tradelines: List[NormalizedTradeline],
              confidence: str = "medium") -> None:
    issues.append(ReviewIssue(
        id=stable_id(issue_type, *[t.id for t in tradelines], label),
        issue_type=issue_type,
        severity=severity,
        customer_label=label,
        customer_explanation=customer,
        admin_explanation=admin,
        suggested_round=round_name,
        related_tradeline_ids=[t.id for t in tradelines],
        evidence=[ev(t) for t in tradelines],
        confidence=confidence,
    ))


def detect_issues(tradelines: List[NormalizedTradeline], groups: List[dict]) -> List[ReviewIssue]:
    issues: List[ReviewIssue] = []

    for t in tradelines:
        blob = " ".join([t.status, t.pay_status, t.remarks, t.raw_block]).lower()

        if any(x in blob for x in COLLECTION_TERMS):
            add_issue(
                issues,
                "collection_review",
                "medium",
                "Collection review",
                "This collection item should be reviewed for original creditor, balance, ownership, and reporting details.",
                "Collection/debt buyer candidate. Verify original creditor, assignment/ownership, balance, authority, and reporting fields.",
                "Round 2 — Collection Review",
                [t],
                t.confidence
            )

        if "charge" in blob and "off" in blob:
            add_issue(
                issues,
                "chargeoff_review",
                "medium",
                "Charge-off review",
                "This charge-off should be reviewed for balance, status, dates, and whether it was sold or transferred.",
                "Charge-off candidate. Check DOFD, balance, sold/transferred status, creditor ownership, and duplicate collection reporting.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                t.confidence
            )

        if not t.date_of_first_delinquency and any(x in blob for x in ["charge", "collection", "delinquent", "past due"]):
            add_issue(
                issues,
                "missing_dofd_review",
                "medium",
                "Important date may be missing",
                "A key delinquency/removal date may need review.",
                "Negative item lacks detected DOFD/removal date. Verify raw report and bureau-specific fields.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                "medium"
            )

        if t.balance and t.balance not in {"$0", "$0.00"} and "closed" in blob and any(x in blob for x in ["transferred", "sold"]):
            add_issue(
                issues,
                "closed_sold_balance_review",
                "medium",
                "Closed or sold account balance review",
                "A closed, sold, or transferred account may still show a balance that should be reviewed.",
                "Closed/sold/transferred item with non-zero balance detected. Needs admin validation.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                "medium"
            )

        if t.confidence == "low":
            add_issue(
                issues,
                "low_confidence_admin_review",
                "low",
                "Needs manual review",
                "This item needs a closer review before any action is recommended.",
                "Parser confidence is low. Admin should verify source snippet and fields.",
                "Admin Review",
                [t],
                "low"
            )

    # Cross-bureau mismatch checks
    by_id = {t.id: t for t in tradelines}
    for group in groups:
        group_items = [by_id[x] for x in group["tradeline_ids"] if x in by_id]
        if len(group_items) < 2:
            continue

        balances = {x.balance for x in group_items if x.balance}
        statuses = {clean_text(x.status or x.pay_status).lower() for x in group_items if x.status or x.pay_status}
        dates_reported = {x.date_reported for x in group_items if x.date_reported}
        dofds = {x.date_of_first_delinquency for x in group_items if x.date_of_first_delinquency}

        if len(balances) >= 2:
            add_issue(
                issues,
                "cross_bureau_balance_mismatch",
                "medium",
                "Balance differs across bureaus",
                "The same or similar account may show different balances across bureaus.",
                "Cross-bureau group has different balances. Verify if mismatch is expected or inaccurate.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

        if len(statuses) >= 2:
            add_issue(
                issues,
                "cross_bureau_status_mismatch",
                "medium",
                "Status differs across bureaus",
                "The same or similar account may show different statuses across bureaus.",
                "Cross-bureau group has different status/pay-status values.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

        if len(dates_reported) >= 2 or len(dofds) >= 2:
            add_issue(
                issues,
                "cross_bureau_date_mismatch",
                "medium",
                "Dates differ across bureaus",
                "The same or similar account may show different important dates across bureaus.",
                "Cross-bureau group has different report dates or DOFD/removal dates.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

    return dedupe_issues(issues)


def dedupe_issues(issues: List[ReviewIssue]) -> List[ReviewIssue]:
    seen = set()
    out = []
    for issue in issues:
        key = (
            issue.issue_type,
            tuple(sorted(issue.related_tradeline_ids)),
            issue.customer_label,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(issue)
    return out


# -----------------------------
# Main parse API
# -----------------------------

def parse_reports(report_texts: Dict[str, dict]) -> ParseResult:
    """
    report_texts format:
    {
      "filename.pdf": {
        "text": "...",
        "bureau": "Experian" optional
      }
    }
    """
    files = []
    all_tradelines: List[NormalizedTradeline] = []

    for filename, payload in report_texts.items():
        text = clean_text(payload.get("text", ""))
        bureau = payload.get("bureau") or detect_bureau(filename, text)
        if bureau == "Unknown Bureau":
            bureau = f"Unknown Report"

        files.append({
            "filename": filename,
            "bureau": bureau,
            "chars": len(text),
            "status": "parsed" if text else "empty_text",
        })

        all_tradelines.extend(parse_tradelines_for_bureau(bureau, filename, text))

    groups = group_cross_bureau(all_tradelines)
    issues = detect_issues(all_tradelines, groups)

    customer_summary = {
        "headline": "Your Credit Check-In was reviewed.",
        "message": "Credit Vivo organized review items into clear categories. No letters or disputes are sent without your approval.",
        "review_items": len(all_tradelines),
        "possible_review_points": len(issues),
        "categories": sorted({i.customer_label for i in issues}),
        "next_step": "Review findings in the dashboard."
    }

    admin_summary = {
        "engine": "Credit Vivo Proprietary Parser Engine",
        "version": "16.0",
        "paid_ai_used": False,
        "tradeline_count": len(all_tradelines),
        "issue_count": len(issues),
        "cross_bureau_group_count": len(groups),
        "warning": "Parser output is draft review data. Verify raw evidence snippets before preparing letters."
    }

    return ParseResult(
        engine="Credit Vivo Proprietary Parser Engine",
        version="16.0",
        paid_ai_used=False,
        files=files,
        tradelines=all_tradelines,
        issues=issues,
        cross_bureau_groups=groups,
        customer_summary=customer_summary,
        admin_summary=admin_summary,
    )


FCRA_NOTICE_OF_DISPUTE = (
    "This is my formal notice of dispute. I dispute the accuracy, completeness, "
    "and/or verifiability of the item identified in this letter. Please conduct a "
    "reasonable investigation under the Fair Credit Reporting Act (FCRA), review all "
    "information I provide with this dispute, forward the dispute and all relevant "
    "information to the furnisher when this is a bureau dispute, and correct, update, "
    "or delete any information that cannot be verified as accurate and complete. "
    "Please mark the item as disputed while the investigation is pending and provide "
    "the written results of the investigation, including any corrected report or "
    "explanation of the verification method used."
)

FCRA_NOTICE_RULES = {
    "consumer_notice_contents": [
        "identify the consumer",
        "identify the account or item being disputed",
        "state the specific information disputed",
        "explain the basis for the dispute",
        "include supporting documents or evidence when available",
        "request correction, update, deletion, or verification results",
    ],
    "bureau_dispute_rules": [
        "consumer reporting agency must conduct a reasonable reinvestigation when accuracy or completeness is disputed",
        "bureau should forward notice of the dispute and relevant information to the furnisher",
        "bureau should provide written reinvestigation results after completion",
        "bureau dispute path is preferred when the next step needs furnisher duties triggered through bureau notice",
    ],
    "furnisher_dispute_rules": [
        "direct furnisher dispute should be sent to the furnisher address shown on the report or other proper direct-dispute address",
        "direct dispute should include enough identifying information, the specific disputed information, the basis for dispute, and supporting evidence",
        "furnisher should conduct a reasonable investigation and review relevant information provided with a proper direct dispute",
        "furnisher should report corrections or stop reporting information that cannot be verified as accurate and complete",
    ],
    "credit_vivo_controls": [
        "do not send automatically",
        "customer approval required before mail, bureau dispute, furnisher dispute, CFPB complaint, state complaint, or attorney escalation",
        "store date sent, recipient, delivery method, tracking number, response due date, response received, and next action",
        "attach evidence packet hash or file reference before sending",
        "record whether the item was marked disputed and whether written results were received",
    ],
}


def _issue_evidence_strength(issue: ReviewIssue) -> str:
    if issue.severity == "high" or issue.confidence == "high":
        return "high"
    if issue.confidence == "low" or issue.severity == "low":
        return "low"
    return "medium"


def _responsible_party_for_issue(issue: ReviewIssue) -> str:
    if issue.issue_type.startswith("cross_bureau"):
        return "bureau_and_furnisher"
    if issue.issue_type in {"collection_review", "chargeoff_review", "closed_sold_balance_review"}:
        return "furnisher_or_collector"
    if issue.issue_type == "missing_dofd_review":
        return "bureau_and_furnisher"
    return "admin_review_required"


def _next_action_for_issue(issue: ReviewIssue) -> str:
    if issue.issue_type.startswith("cross_bureau"):
        return "round_2_field_level_bureau_dispute"
    if issue.issue_type in {"collection_review", "chargeoff_review", "closed_sold_balance_review"}:
        return "furnisher_direct_dispute_after_bureau_review"
    if issue.issue_type == "missing_dofd_review":
        return "round_2_bureau_dispute_then_reinvestigation_if_unverified"
    return "admin_review_before_letter"


def _draft_letter_subject(letter_type: str, issue: ReviewIssue) -> str:
    if letter_type == "furnisher_direct_dispute":
        return f"Direct Dispute of Account Reporting - {issue.customer_label}"
    if letter_type == "bureau_dispute":
        return f"Credit Report Dispute - {issue.customer_label}"
    return f"Admin Review Required - {issue.customer_label}"


def _draft_letter_body(letter_type: str, recipient_type: str, issue: ReviewIssue) -> str:
    if letter_type == "admin_review_hold":
        return (
            "DRAFT HOLD - ADMIN REVIEW REQUIRED\n\n"
            f"Issue: {issue.customer_label}\n"
            f"Reason: {issue.customer_explanation}\n\n"
            "Credit Vivo did not generate a send-ready letter for this item because it needs "
            "manual review before any dispute path is selected."
        )

    recipient_line = "Credit Bureau" if recipient_type == "credit_bureau" else "Furnisher / Collector"
    requested_action = (
        "Please investigate this item, correct any inaccurate or incomplete information, "
        "delete any information that cannot be verified, and send the investigation results in writing."
        if letter_type == "bureau_dispute"
        else
        "Please provide the basis for your reporting, including records supporting ownership, "
        "balance, status, payment history, date of first delinquency, and authority to report this account."
    )
    evidence_note = "Evidence from the Credit Vivo scanner is attached for customer/admin review."
    if issue.evidence:
        snippet = issue.evidence[0].snippet[:450]
        evidence_note = f"Scanner evidence excerpt for review: {snippet}"

    return (
        "DRAFT - CUSTOMER REVIEW AND APPROVAL REQUIRED\n"
        "DO NOT SEND UNTIL CUSTOMER AUTHORIZATION IS VERIFIED\n\n"
        "[Customer Name]\n"
        "[Customer Mailing Address]\n"
        "[City, State ZIP]\n\n"
        "[Date]\n\n"
        f"To: {recipient_line}\n"
        "[Recipient Address]\n\n"
        f"Re: {issue.customer_label}\n\n"
        "To Whom It May Concern:\n\n"
        "I am disputing the accuracy, completeness, or verifiability of the credit reporting item "
        "identified below. Please treat this letter as my formal notice of dispute.\n\n"
        f"Disputed issue: {issue.customer_label}\n"
        f"Reason for dispute: {issue.customer_explanation}\n"
        f"Review round: {issue.suggested_round}\n\n"
        f"{FCRA_NOTICE_OF_DISPUTE}\n\n"
        f"{requested_action}\n\n"
        f"{evidence_note}\n\n"
        "Please send your written response to the mailing address above.\n\n"
        "Sincerely,\n\n"
        "[Customer Signature]\n"
        "[Customer Printed Name]\n"
    )


def build_letter_workflow() -> dict:
    return {
        "draft_only": True,
        "send_letters_automatically": False,
        "customer_authorization_required": True,
        "fcra_notice_of_dispute": FCRA_NOTICE_OF_DISPUTE,
        "fcra_notice_rules": FCRA_NOTICE_RULES,
        "bureau_dispute_procedure": {
            "recipient_type": "credit_bureau",
            "delivery_preference": "certified_mail_for_important_disputes",
            "round_1_uses": [
                "wrong balance or status",
                "unrecognized account",
                "duplicate collection",
                "original creditor and debt buyer both showing active balance",
                "obsolete or missing key date item",
                "obvious mismatch across bureaus",
            ],
            "packet_checklist": [
                "customer-approved dispute letter",
                "FCRA notice of dispute",
                "specific disputed item and reason",
                "targeted proof only",
                "ID and proof of address when needed",
                "redacted unrelated account data",
                "highlighted disputed item",
                "request written investigation results",
            ],
        },
        "furnisher_direct_dispute_procedure": {
            "recipient_type": "furnisher_or_collector",
            "delivery_preference": "certified_mail",
            "prerequisites": [
                "consumer-specific issue identified",
                "customer authorization verified",
                "evidence packet reviewed by admin",
                "specific furnisher item and dispute reason stated",
                "proper furnisher/direct-dispute address confirmed when available",
            ],
            "requested_verification": [
                "basis for reporting",
                "contract or account records",
                "itemized balance",
                "chain of title or assignment where applicable",
                "payment history",
                "date of first delinquency support",
                "proof reporting is complete and accurate",
                "written investigation result or correction/deletion notice",
            ],
        },
        "escalation_procedure": {
            "cfpb_complaint": {
                "trigger": "no response, verified-as-accurate with weak support, or repeated inaccurate reporting after dispute history is complete",
                "packet": "original dispute, proof of delivery, bureau/furnisher responses, current report excerpt, damages or denial evidence if available",
            },
            "state_attorney_general": {
                "trigger": "pattern of non-response, abusive collection conduct, or unresolved state-law concern",
                "packet": "same evidence packet plus consumer timeline and requested resolution",
            },
            "state_regulator": {
                "trigger": "licensed collector, lender, or credit repair/regulatory issue needs agency review",
                "packet": "collector/furnisher identity, license information if known, dispute timeline, and supporting evidence",
            },
            "attorney_review": {
                "trigger": "strong FCRA/FDCPA pattern, measurable damages, denial letter, identity theft, mixed file, or repeated verification without reasonable investigation",
                "packet": "full dispute history, reports before and after disputes, notices, responses, delivery proofs, and damages evidence",
            },
        },
        "tracking_schema": [
            "letter_id",
            "issue_ids",
            "recipient_type",
            "recipient_name",
            "recipient_address",
            "delivery_method",
            "certified_tracking_number",
            "sent_date",
            "delivered_date",
            "response_due_date",
            "day_15_check_date",
            "day_35_followup_check_date",
            "response_received_date",
            "current_status",
            "next_action",
            "fcra_notice_included",
            "customer_authorization_verified",
            "evidence_packet_hash",
        ],
        "event_log_entries": [
            "letter_drafted",
            "customer_authorization_verified",
            "fcra_notice_added",
            "proof_packet_attached",
            "certified_mail_queued",
            "tracking_number_saved",
            "delivery_confirmed",
            "response_deadline_created",
            "response_received",
            "response_scanned",
            "next_action_assigned",
        ],
    }


def build_recommended_letter_queue(issues: List[ReviewIssue]) -> List[dict]:
    queue = []
    for issue in issues:
        responsible_party = _responsible_party_for_issue(issue)
        letter_type = "bureau_dispute"
        recipient_type = "credit_bureau"
        if responsible_party == "furnisher_or_collector":
            letter_type = "furnisher_direct_dispute"
            recipient_type = "furnisher_or_collector"
        elif responsible_party == "admin_review_required":
            letter_type = "admin_review_hold"
            recipient_type = "undetermined"

        queue.append({
            "letter_id": stable_id("letter", issue.id, letter_type),
            "issue_id": issue.id,
            "issue_type": issue.issue_type,
            "letter_type": letter_type,
            "letter_subject": _draft_letter_subject(letter_type, issue),
            "draft_letter_body": _draft_letter_body(letter_type, recipient_type, issue),
            "round": issue.suggested_round,
            "recipient_type": recipient_type,
            "responsible_party": responsible_party,
            "delivery_method": "certified_mail_recommended",
            "fcra_notice_required": letter_type != "admin_review_hold",
            "fcra_notice_included": letter_type != "admin_review_hold",
            "customer_approval_required": True,
            "customer_authorization_verified": False,
            "tracking_status": "draft_not_sent",
            "recommended_next_action": _next_action_for_issue(issue),
            "escalation_candidate": False,
        })
    return queue


def build_fcra_review(issues: List[ReviewIssue]) -> List[dict]:
    return [
        {
            "issue_id": issue.id,
            "possible_fcra_issue": issue.issue_type != "low_confidence_admin_review",
            "issue_type": issue.issue_type,
            "responsible_party": _responsible_party_for_issue(issue),
            "dispute_history_complete": False,
            "evidence_strength": _issue_evidence_strength(issue),
            "damages_evidence": "none",
            "next_action": _next_action_for_issue(issue),
            "requires_admin_review": True,
        }
        for issue in issues
    ]


METRO2_FCRA_RULES = {
    "collection_review": {
        "metro2_fields": ["Account Type", "Portfolio Type", "Original Creditor", "Current Balance", "Account Status", "Date Reported"],
        "fcra_focus": "Accuracy, completeness, ownership, and verification of collection reporting.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["original creditor", "assignment or ownership proof", "balance calculation", "reporting authority", "account-level source records"],
        "dispute_theory": "The collection must be accurate, complete, and verifiable. If ownership, balance, or reporting authority cannot be supported, correction or deletion may be appropriate.",
    },
    "chargeoff_review": {
        "metro2_fields": ["Account Status", "Payment Rating", "Current Balance", "Amount Past Due", "Date of First Delinquency", "Date Closed", "Special Comment"],
        "fcra_focus": "Charge-off reporting accuracy, dates, balance, sold/transferred status, and obsolescence risk.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice", "FCRA 605 obsolescence review"],
        "evidence_needed": ["charge-off ledger", "payment history", "DOFD support", "sale or transfer record", "balance breakdown"],
        "dispute_theory": "A charge-off should not report internally inconsistent status, balance, ownership, or delinquency dates.",
    },
    "missing_dofd_review": {
        "metro2_fields": ["Date of First Delinquency", "FCRA Compliance/Date of First Delinquency", "Estimated Removal Date"],
        "fcra_focus": "Missing or unsupported delinquency date for negative reporting.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 605 obsolescence review", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["first delinquency date", "payment history", "charge-off/collection timeline", "bureau reporting history"],
        "dispute_theory": "Negative reporting needs a supportable delinquency timeline so the account is not reported longer than allowed.",
    },
    "closed_sold_balance_review": {
        "metro2_fields": ["Account Status", "Current Balance", "Amount Past Due", "Date Closed", "Special Comment", "Purchased/Sold/Transferred Indicator"],
        "fcra_focus": "Closed, sold, transferred, or assigned account still reporting a balance.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["sale/transfer record", "current owner", "balance ledger", "zero-balance update support"],
        "dispute_theory": "If the account was sold or transferred, the reporting balance/status must match the furnisher's actual ownership and records.",
    },
    "cross_bureau_balance_mismatch": {
        "metro2_fields": ["Current Balance", "Amount Past Due", "High Credit/Original Amount", "Date Reported"],
        "fcra_focus": "Same account reports different balances across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["current account ledger", "last statement", "bureau reporting snapshots", "furnisher update history"],
        "dispute_theory": "The same account should have a supportable balance. Differences may be explainable by timing, but unsupported differences should be corrected.",
    },
    "cross_bureau_status_mismatch": {
        "metro2_fields": ["Account Status", "Payment Rating", "Special Comment", "Compliance Condition Code"],
        "fcra_focus": "Same account reports different status/payment condition across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["account status records", "payment history", "charge-off/collection status support", "bureau update records"],
        "dispute_theory": "Status and payment condition should be consistent with the furnisher's records and not materially misleading.",
    },
    "cross_bureau_date_mismatch": {
        "metro2_fields": ["Date Opened", "Date Closed", "Date Reported", "Date of First Delinquency", "Estimated Removal Date"],
        "fcra_focus": "Same account reports different key dates across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 605 obsolescence review", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["opening records", "closing records", "DOFD support", "payment history", "bureau report snapshots"],
        "dispute_theory": "Key dates drive legal reporting age and consumer harm. Unsupported date differences should be corrected or deleted.",
    },
    "low_confidence_admin_review": {
        "metro2_fields": ["Raw tradeline block", "All extracted fields"],
        "fcra_focus": "Manual validation before dispute strategy.",
        "fcra_sections": ["Admin review required before FCRA dispute theory is selected"],
        "evidence_needed": ["raw report excerpt", "account-level PDF pages", "manual field validation"],
        "dispute_theory": "Do not send a dispute until the extracted fields are confirmed against the raw report.",
    },
}


def _expert_rule_for_issue(issue: ReviewIssue) -> dict:
    return METRO2_FCRA_RULES.get(issue.issue_type, {
        "metro2_fields": ["Account Status", "Current Balance", "Date Reported", "Remarks"],
        "fcra_focus": "Accuracy, completeness, and verifiability review.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["raw report excerpt", "furnisher records", "bureau investigation response"],
        "dispute_theory": "If the reporting cannot be verified as accurate and complete, it should be corrected, updated, or deleted.",
    })


def build_metro2_fcra_review(issues: List[ReviewIssue]) -> List[dict]:
    review = []
    for issue in issues:
        rule = _expert_rule_for_issue(issue)
        review.append({
            "issue_id": issue.id,
            "issue_type": issue.issue_type,
            "customer_label": issue.customer_label,
            "severity": issue.severity,
            "confidence": issue.confidence,
            "metro2_fields_to_review": rule["metro2_fields"],
            "fcra_focus": rule["fcra_focus"],
            "fcra_sections": rule["fcra_sections"],
            "evidence_needed": rule["evidence_needed"],
            "dispute_theory": rule["dispute_theory"],
            "responsible_party": _responsible_party_for_issue(issue),
            "recommended_next_action": _next_action_for_issue(issue),
            "customer_approval_required": True,
            "attorney_review_signal": issue.severity == "high" or issue.issue_type.startswith("cross_bureau"),
        })
    return review


def result_to_dict(result: ParseResult) -> dict:
    return {
        "engine": result.engine,
        "version": result.version,
        "paid_ai_used": result.paid_ai_used,
        "files": result.files,
        "tradelines": [asdict(t) for t in result.tradelines],
        "issues": [asdict(i) for i in result.issues],
        "cross_bureau_groups": result.cross_bureau_groups,
        "customer_summary": result.customer_summary,
        "admin_summary": result.admin_summary,
        "letter_workflow": build_letter_workflow(),
        "recommended_letter_queue": build_recommended_letter_queue(result.issues),
        "fcra_review": build_fcra_review(result.issues),
        "metro2_fcra_review": build_metro2_fcra_review(result.issues),
    }


def _safe_workbook_cell(value):
    if isinstance(value, (list, tuple)):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return value


def _write_workbook_sheet(sheet, rows: List[List[object]]) -> None:
    for row in rows:
        sheet.append([_safe_workbook_cell(value) for value in row])

    if not rows or Workbook is None:
        return

    header_fill = PatternFill("solid", fgColor="D1FAE5")
    header_font = Font(bold=True, color="064E3B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, column in enumerate(sheet.columns, start=1):
        max_length = 10
        for cell in column:
            max_length = max(max_length, min(len(str(cell.value or "")), 55))
        sheet.column_dimensions[get_column_letter(column_index)].width = max_length + 2

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _norm_compare(value) -> str:
    return clean_text(str(value or "")).lower()


def _comparison_flag(values: List[str], label: str) -> str:
    present = {_norm_compare(value) for value in values if _norm_compare(value)}
    if len(present) >= 2:
        return f"{label} differs"
    return ""


def build_three_bureau_comparison_rows(data: dict) -> List[List[object]]:
    tradelines_by_id = {item.get("id"): item for item in data.get("tradelines", [])}
    bureau_order = ["Equifax", "Experian", "TransUnion"]
    cross_issue_ids = set()
    per_bureau_fields = [
        ("Source", "source_filename"),
        ("Account #", "account_number_masked"),
        ("Type", "account_type"),
        ("Original Creditor", "original_creditor"),
        ("Balance", "balance"),
        ("Past Due", "past_due"),
        ("Status", "status"),
        ("Opened", "date_opened"),
        ("Closed", "date_closed"),
        ("Reported", "date_reported"),
        ("DOFD", "date_of_first_delinquency"),
        ("Remarks", "remarks"),
        ("Confidence", "confidence"),
        ("Raw Evidence", "raw_block"),
    ]

    for issue in data.get("issues", []):
        if str(issue.get("issue_type", "")).startswith("cross_bureau"):
            cross_issue_ids.update(issue.get("related_tradeline_ids", []))

    rows = [["Account Name"]]
    for bureau in bureau_order:
        rows[0].extend([f"{bureau} {label}" for label, _field in per_bureau_fields])
    rows[0].extend([
        "Errors",
        "Findings",
        "Dispute Targets",
        "Bureau Dispute Draft",
        "Furnisher Dispute Draft",
        "Tracking Status",
        "Missing Bureaus",
        "Matched Bureaus",
        "Group ID",
    ])

    for group in data.get("cross_bureau_groups", []):
        items = [tradelines_by_id.get(item_id) for item_id in group.get("tradeline_ids", [])]
        items = [item for item in items if item]
        if len(items) < 2:
            continue

        by_bureau = {item.get("bureau"): item for item in items}
        balances = [item.get("balance", "") for item in items]
        statuses = [item.get("status") or item.get("pay_status") or "" for item in items]
        reported_dates = [item.get("date_reported", "") for item in items]
        dofds = [item.get("date_of_first_delinquency", "") for item in items]
        account_names = [item.get("account_name", "") for item in items]

        flags = [
            _comparison_flag(balances, "Balance"),
            _comparison_flag(statuses, "Status"),
            _comparison_flag(reported_dates, "Reported date"),
            _comparison_flag(dofds, "DOFD"),
            _comparison_flag(account_names, "Account name"),
        ]
        missing = [bureau for bureau in bureau_order if bureau not in by_bureau]
        if missing:
            flags.append("Missing on " + ", ".join(missing))

        has_cross_issue = any(item.get("id") in cross_issue_ids for item in items)
        suggested_review = (
            "Review and dispute field-level mismatch if inaccurate or unverifiable."
            if has_cross_issue or any(flags)
            else
            "Matched across bureaus. No mismatch flag detected."
        )
        account_name = "; ".join(sorted({name for name in account_names if name}))
        error_text = "; ".join(flag for flag in flags if flag) or "No mismatch flag detected."
        dispute_targets = "Credit bureaus and furnishing creditor/collector"
        bureau_letter = (
            "DRAFT - CUSTOMER REVIEW AND APPROVAL REQUIRED\n"
            "To: Credit Bureau\n"
            f"Re: {account_name}\n\n"
            "I dispute the accuracy, completeness, and/or verifiability of this account as reported on my credit file. "
            f"The bureau comparison shows: {error_text}. "
            "Please conduct a reasonable FCRA investigation, forward all relevant dispute information to the furnisher, "
            "mark the item as disputed while pending, and correct, update, or delete any information that cannot be verified as accurate and complete. "
            "Please provide written investigation results and an updated report if changes are made."
        )
        furnisher_letter = (
            "DRAFT - CUSTOMER REVIEW AND APPROVAL REQUIRED\n"
            "To: Furnisher / Collector\n"
            f"Re: {account_name}\n\n"
            "I dispute the accuracy, completeness, and/or verifiability of the information you are furnishing about this account. "
            f"The bureau comparison shows: {error_text}. "
            "Please investigate your records and provide the basis for reporting, including balance support, account status support, "
            "date reporting support, ownership/assignment records where applicable, and any records used to verify the account. "
            "If the information cannot be verified as accurate and complete, please correct, update, or stop furnishing the disputed information."
        )

        row = [account_name]

        for bureau in bureau_order:
            item = by_bureau.get(bureau, {})
            for _label, field in per_bureau_fields:
                value = item.get(field, "")
                if field == "status":
                    value = item.get("status") or item.get("pay_status") or ""
                if field == "raw_block":
                    value = clean_text(str(value))[:650]
                row.append(value)
        row.extend([
            error_text,
            suggested_review,
            dispute_targets,
            bureau_letter,
            furnisher_letter,
            "draft_not_sent_customer_approval_required",
            ", ".join(missing),
            ", ".join(sorted(by_bureau.keys())),
            group.get("group_id", ""),
        ])
        rows.append(row)

    if len(rows) == 1:
        rows.append([
            "No matched cross-bureau accounts",
            *["" for _ in range(len(bureau_order) * len(per_bureau_fields))],
            "No 3-bureau comparison could be created from the uploaded report set.",
            "Upload reports from at least two bureaus to compare the same account side by side.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    return rows


def write_desktop_workbook(data: dict, out_dir: Path) -> None:
    if Workbook is None:
        return

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    bureau_comparison = wb.create_sheet("3 Bureau Comparison")
    errors = wb.create_sheet("Detected Errors")
    items = wb.create_sheet("Review Items")
    metro2_fcra = wb.create_sheet("Metro 2 + FCRA Review")
    fcra_notice_rules = wb.create_sheet("FCRA Notice Rules")
    letters = wb.create_sheet("Draft Letters")
    fcra = wb.create_sheet("FCRA Review")

    _write_workbook_sheet(summary, [
        ["Credit Vivo Scanner Output", ""],
        ["Engine", data.get("engine", "")],
        ["Version", data.get("version", "")],
        ["Paid AI Used", "Yes" if data.get("paid_ai_used") else "No"],
        ["Files Parsed", len(data.get("files", []))],
        ["Review Items", len(data.get("tradelines", []))],
        ["Detected Errors / Review Points", len(data.get("issues", []))],
        ["Draft Letters Queued", len(data.get("recommended_letter_queue", []))],
        ["Customer Message", data.get("customer_summary", {}).get("message", "")],
        ["Important Notice", "Draft review data only. Nothing is sent without customer approval and admin review."],
    ])

    _write_workbook_sheet(bureau_comparison, build_three_bureau_comparison_rows(data))

    _write_workbook_sheet(errors, [
        ["Issue ID", "Issue Type", "Severity", "Customer Label", "Customer Explanation", "Admin Explanation", "Suggested Round", "Related Tradeline IDs", "Confidence", "Evidence Count"],
        *[
            [
                issue.get("id", ""),
                issue.get("issue_type", ""),
                issue.get("severity", ""),
                issue.get("customer_label", ""),
                issue.get("customer_explanation", ""),
                issue.get("admin_explanation", ""),
                issue.get("suggested_round", ""),
                issue.get("related_tradeline_ids", []),
                issue.get("confidence", ""),
                len(issue.get("evidence", [])),
            ]
            for issue in data.get("issues", [])
        ],
    ])

    _write_workbook_sheet(items, [
        ["ID", "Bureau", "Source File", "Account Name", "Account Type", "Status", "Balance", "Date Opened", "Date Reported", "DOFD", "Remarks", "Confidence", "Needs Admin Review"],
        *[
            [
                item.get("id", ""),
                item.get("bureau", ""),
                item.get("source_filename", ""),
                item.get("account_name", ""),
                item.get("account_type", ""),
                item.get("status", ""),
                item.get("balance", ""),
                item.get("date_opened", ""),
                item.get("date_reported", ""),
                item.get("date_of_first_delinquency", ""),
                item.get("remarks", ""),
                item.get("confidence", ""),
                "Yes" if item.get("needs_admin_review") else "No",
            ]
            for item in data.get("tradelines", [])
        ],
    ])

    _write_workbook_sheet(metro2_fcra, [
        [
            "Issue ID",
            "Issue Type",
            "Customer Label",
            "Metro 2 Fields To Review",
            "FCRA Focus",
            "FCRA Sections / Duties",
            "Evidence Needed",
            "Dispute Theory",
            "Responsible Party",
            "Recommended Next Action",
            "Attorney Review Signal",
            "Customer Approval Required",
        ],
        *[
            [
                row.get("issue_id", ""),
                row.get("issue_type", ""),
                row.get("customer_label", ""),
                row.get("metro2_fields_to_review", []),
                row.get("fcra_focus", ""),
                row.get("fcra_sections", []),
                row.get("evidence_needed", []),
                row.get("dispute_theory", ""),
                row.get("responsible_party", ""),
                row.get("recommended_next_action", ""),
                "Yes" if row.get("attorney_review_signal") else "No",
                "Yes" if row.get("customer_approval_required") else "No",
            ]
            for row in data.get("metro2_fcra_review", [])
        ],
    ])

    notice_rows = [["Rule Area", "Requirement / Control"]]
    for area, controls in data.get("letter_workflow", {}).get("fcra_notice_rules", {}).items():
        label = area.replace("_", " ").title()
        for control in controls:
            notice_rows.append([label, control])
    _write_workbook_sheet(fcra_notice_rules, notice_rows)

    _write_workbook_sheet(letters, [
        ["Letter ID", "Issue ID", "Subject", "Letter Type", "Round", "Recipient Type", "Delivery Method", "FCRA Notice Included", "Customer Approval Required", "Tracking Status", "Recommended Next Action", "Draft Letter Body"],
        *[
            [
                letter.get("letter_id", ""),
                letter.get("issue_id", ""),
                letter.get("letter_subject", ""),
                letter.get("letter_type", ""),
                letter.get("round", ""),
                letter.get("recipient_type", ""),
                letter.get("delivery_method", ""),
                "Yes" if letter.get("fcra_notice_included") else "No",
                "Yes" if letter.get("customer_approval_required") else "No",
                letter.get("tracking_status", ""),
                letter.get("recommended_next_action", ""),
                letter.get("draft_letter_body", ""),
            ]
            for letter in data.get("recommended_letter_queue", [])
        ],
    ])

    _write_workbook_sheet(fcra, [
        ["Issue ID", "Possible FCRA Issue", "Issue Type", "Responsible Party", "Dispute History Complete", "Evidence Strength", "Damages Evidence", "Next Action", "Requires Admin Review"],
        *[
            [
                row.get("issue_id", ""),
                "Yes" if row.get("possible_fcra_issue") else "No",
                row.get("issue_type", ""),
                row.get("responsible_party", ""),
                "Yes" if row.get("dispute_history_complete") else "No",
                row.get("evidence_strength", ""),
                row.get("damages_evidence", ""),
                row.get("next_action", ""),
                "Yes" if row.get("requires_admin_review") else "No",
            ]
            for row in data.get("fcra_review", [])
        ],
    ])

    wb.save(out_dir / "credit_vivo_desktop_scanner_output.xlsx")


def write_outputs(result: ParseResult, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    data = result_to_dict(result)
    (out_dir / "credit_vivo_parser_result.json").write_text(json.dumps(data, indent=2), encoding="utf-8")

    # Tradelines CSV
    tradeline_rows = [asdict(t) for t in result.tradelines]
    if tradeline_rows:
        with (out_dir / "tradelines.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(tradeline_rows[0].keys()))
            writer.writeheader()
            writer.writerows(tradeline_rows)

    # Issues CSV - flatten evidence count only
    issue_rows = []
    for issue in result.issues:
        d = asdict(issue)
        d["related_tradeline_ids"] = ";".join(issue.related_tradeline_ids)
        d["evidence_count"] = len(issue.evidence)
        d.pop("evidence", None)
        issue_rows.append(d)

    if issue_rows:
        with (out_dir / "review_issues.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(issue_rows[0].keys()))
            writer.writeheader()
            writer.writerows(issue_rows)

    letter_sections = []
    for letter in data.get("recommended_letter_queue", []):
        letter_sections.append(
            "\n".join(
                [
                    f"Letter ID: {letter.get('letter_id', '')}",
                    f"Subject: {letter.get('letter_subject', '')}",
                    f"Type: {letter.get('letter_type', '')}",
                    f"Round: {letter.get('round', '')}",
                    f"Recipient: {letter.get('recipient_type', '')}",
                    f"Tracking status: {letter.get('tracking_status', '')}",
                    "",
                    str(letter.get("draft_letter_body") or "No draft body generated."),
                ]
            )
        )

    if letter_sections:
        (out_dir / "draft_dispute_letters.txt").write_text(
            ("\n\n" + ("-" * 72) + "\n\n").join(letter_sections),
            encoding="utf-8",
        )

    write_desktop_workbook(data, out_dir)


# Phase 3 draft-only integrations.
# These classes intentionally do not send disputes, mail, or bank data by
# themselves. They prepare review artifacts for an authenticated admin workflow.
import hashlib
import os
import random
import re
from datetime import datetime, timezone
from typing import Any, Optional

import requests
from sodapy import Socrata


class OpenDataCrossMatcher:
    """Cross-match collector names against public licensing data."""

    def __init__(self, domain: str = "opendata.maryland.gov"):
        self.domain = domain
        self.app_token = os.getenv("SOCRATA_APP_TOKEN")
        self.client = Socrata(self.domain, self.app_token or None)

    @staticmethod
    def _safe_collector_name(collector_name: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9 '&.,-]", "", collector_name or "").strip()
        return cleaned[:120]

    def verify_collector_license(self, collector_name: str, state_code: str = "MD") -> dict[str, Any]:
        safe_name = self._safe_collector_name(collector_name)
        safe_state = (state_code or "MD").upper()[:2]

        if not safe_name:
            return {
                "found": False,
                "status": "UNKNOWN",
                "leverage_flag": None,
                "review_note": "Collector name was empty or unsupported.",
            }

        try:
            dataset_id = "gdzy-2fen"
            escaped_name = safe_name.upper().replace("'", "''")
            query = f"business_name like '%{escaped_name}%'"
            results = self.client.get(dataset_id, where=query, limit=1)

            if not results:
                return {
                    "found": False,
                    "status": "UNKNOWN",
                    "leverage_flag": None,
                    "review_note": "No public licensing record matched this collector name.",
                    "last_checked_at": datetime.now(timezone.utc).isoformat(),
                }

            record = results[0]
            status = str(record.get("license_status", "UNKNOWN")).upper()
            inactive_statuses = {"REVOKED", "EXPIRED", "SUSPENDED", "INACTIVE"}
            leverage_flag: Optional[str] = None

            if status in inactive_statuses:
                leverage_flag = (
                    f"Public open-data records may indicate {safe_name} has a {status} "
                    f"license status in {safe_state}. Admin review should verify the "
                    "source record before using this in a customer letter."
                )

            return {
                "found": True,
                "status": status,
                "leverage_flag": leverage_flag,
                "raw_evidence": record,
                "last_checked_at": datetime.now(timezone.utc).isoformat(),
            }
        except Exception as exc:
            return {
                "found": False,
                "status": "ERROR",
                "leverage_flag": None,
                "error": str(exc),
                "last_checked_at": datetime.now(timezone.utc).isoformat(),
            }


class EOSCARBypassSpinner:
    """Generate unique, draft-only dispute letter language for admin review."""

    def __init__(self):
        self.openers = [
            "I am requesting an investigation of an item appearing on my {bureau} credit file.",
            "Please review my {bureau} credit profile, including the tradeline reported by {creditor}.",
            "I am asking for a reasonable reinvestigation of the {creditor} account on my {bureau} report.",
        ]
        self.closers = [
            "Please complete a reasonable investigation and provide the results in writing.",
            "Please verify the reporting with competent evidence or correct the information as required.",
            "Please send the investigation results and any updated report after your review is complete.",
        ]

    @staticmethod
    def _mask_account(account_num: str) -> str:
        digits = re.sub(r"\D", "", account_num or "")
        if len(digits) <= 4:
            return "ending in " + (digits or "unknown")
        return "ending in " + digits[-4:]

    def generate_unique_letter(
        self,
        bureau: str,
        creditor: str,
        account_num: str,
        violation_code: str,
        open_data_leverage: str | None = None,
    ) -> str:
        safe_bureau = bureau or "the bureau"
        safe_creditor = creditor or "the furnisher"
        safe_violation = violation_code or "reporting accuracy review"
        opener = random.choice(self.openers).format(
            bureau=safe_bureau,
            creditor=safe_creditor,
            account_num=self._mask_account(account_num),
        )
        closer = random.choice(self.closers)
        body = (
            f"This draft concerns {safe_creditor}, account {self._mask_account(account_num)}. "
            f"The item is being reviewed for {safe_violation}. I am not asking for any "
            "accurate and verifiable information to be removed; I am asking that the "
            "reporting be investigated, corrected, updated, or deleted only if it cannot "
            "be verified under applicable law."
        )

        if open_data_leverage:
            body += (
                f"\n\nAdditional review note: {open_data_leverage} This public-data signal "
                "must be verified by an admin before the language is sent."
            )

        return f"{opener}\n\n{body}\n\n{closer}\n\nDraft only - requires customer approval and admin review."


class LitigationTracker:
    """Prepare mail-tracking records without transmitting mail automatically."""

    def __init__(self):
        self.lob_api_key = os.getenv("LOB_API_KEY")
        self.lob_base_url = os.getenv("LOB_BASE_URL", "https://api.lob.com/v1")

    @staticmethod
    def hash_tracking_number(tracking_number: str) -> str:
        normalized = re.sub(r"\s+", "", tracking_number or "").upper()
        if not normalized:
            raise ValueError("tracking_number is required")
        return hashlib.sha256(normalized.encode("utf-8")).hexdigest()

    def build_mail_tracking_event(
        self,
        letter_id: str,
        tracking_number: str,
        delivery_status: str,
        delivery_timestamp: str | None = None,
    ) -> dict[str, Any]:
        return {
            "letter_id": letter_id,
            "usps_tracking_hash": self.hash_tracking_number(tracking_number),
            "delivery_status": (delivery_status or "UNKNOWN").upper(),
            "delivery_timestamp": delivery_timestamp,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

    def fetch_lob_letter_status(self, lob_letter_id: str) -> dict[str, Any]:
        if not self.lob_api_key:
            return {
                "ok": False,
                "error": "LOB_API_KEY is not configured.",
                "review_note": "Mail status lookup is disabled in local testing.",
            }

        response = requests.get(
            f"{self.lob_base_url}/letters/{lob_letter_id}",
            auth=(self.lob_api_key, ""),
            timeout=15,
        )
        response.raise_for_status()
        payload = response.json()
        return {
            "ok": True,
            "letter_id": payload.get("id"),
            "mail_type": payload.get("mail_type"),
            "expected_delivery_date": payload.get("expected_delivery_date"),
            "tracking_events": payload.get("tracking_events", []),
        }
